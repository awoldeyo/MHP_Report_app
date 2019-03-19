from pandas import DataFrame, to_datetime, ExcelWriter

from jira.client import JIRA
from jira.exceptions import JIRAError

from openpyxl import load_workbook


class Connection(object):
    def __init__(self, username, password):
        self.url = 'http://jira.mymhp.net/'
        self.username = username
        self.password = password

    def authenticate(self):
        try:
            self.jira_options={'server': self.url}
            self.jira=JIRA(options=self.jira_options,
                           basic_auth=(f'{self.username}', f'{self.password}'))
            return f'You are logged in as {self.jira.current_user()}!'
        except JIRAError as j:
            return f'{j}'

    def generate_opl(self):
        issues = self.jira.search_issues(
                jql_str='project = DSGVOOPL AND labels = OPL',
                maxResults=False)
        table = []
        for issue in issues:
            row = {}
            row['Nr.'] = issue.key
            row['Themenschwerpunkt'] = ", ".join([c.name for c in issue.fields.components if c is not None])
            row['Bezeichnung'] = issue.fields.summary

            # Get comments
            comment = self.jira.comments(issue)
            try:
                row['Bearbeitungsstand'] = "\n".join([c.body for c in comment])
            except IndexError as i:
                print(f"No comment available for {issue.key}")
                row['Bearbeitungsstand'] = None

            row['Erstelldatum'] = to_datetime(issue.fields.created)
            try:
                row['Zieldatum'] = to_datetime(issue.fields.duedate)
            except AttributeError:
                row['Zieldatum'] = issue.fields.duedate

            row['Status'] = issue.fields.status.name
            try:
                row['Verantwortung im Projekt'] = ", ".join(issue.fields.customfield_11008)
            except TypeError:
                row['Verantwortung im Projekt'] = issue.fields.customfield_11008

            try:
                row['Verantwortung der Umsetzung'] = ", ".join(issue.fields.customfield_11009)
            except TypeError:
                row['Verantwortung der Umsetzung'] = issue.fields.customfield_11009
            table.append(row)

        df = DataFrame(table)
        # Sort dataframe columns
        df = df.reindex(labels=['Nr.',
                                'Themenschwerpunkt',
                                'Bezeichnung',
                                'Bearbeitungsstand',
                                'Erstelldatum',
                                'Zieldatum',
                                'Status',
                                'Verantwortung der Umsetzung',
                                'Verantwortung im Projekt',
                               ],
                        axis=1)

        # Clean up Status categories
        status_mapping = {'Fertig': 'Erledigt',
                          'Backlog': 'Offen',
                          'Selected for Development':'Offen'}
        df['Status'] = df.Status.transform(lambda x: status_mapping.get(x,x))

        return df
    def generate_risk(self):
        issues = self.jira.search_issues(
                jql_str='project = DSGVOOPL AND labels = Risikoliste AND status != Done',
                maxResults=False)
        table = []
        for issue in issues:
            row = {}
            row['Nr.'] = issue.key
            row['Themenschwerpunkt'] = ", ".join([c.name for c in issue.fields.components if c is not None])
            row['Risikobezeichnung'] = issue.fields.summary
            row['Beschreibung'] = issue.fields.description


            status_mapping = {'Fertig': 'Erledigt',
                              'Backlog': 'Offen',
                              'Selected for Development':'Offen'}
            comment = ["Siehe OPL:"]
            linked = []
            for linkedissue in issue.fields.issuelinks:
                try:
                    key = linkedissue.outwardIssue.key
                except:
                    key = linkedissue.inwardIssue.key
                status = status_mapping.get(self.jira.issue(key).fields.status.name)
                statement = f"{key} (Status: {status})"
                linked.append(statement)
            linked = ", ".join(linked)
            comment.append(linked)
            if linked != '':
                row['Bearbeitungsstand'] = " ".join(comment)
            else:
                row['Bearbeitungsstand'] = ""

            try:
                row['Risikoart'] = issue.fields.customfield_11006.value
            except:
                row['Risikoart'] = None

            row['Erstelldatum'] = to_datetime(issue.fields.created)

            row['Risikoeinstufung'] = issue.fields.priority.name

            try:
                row['Verantwortung im Projekt'] = ", ".join(issue.fields.customfield_11008)
            except TypeError:
                row['Verantwortung im Projekt'] = issue.fields.customfield_11008

            try:
                row['Verantwortung der Umsetzung'] = ", ".join(issue.fields.customfield_11009)
            except TypeError:
                row['Verantwortung der Umsetzung'] = issue.fields.customfield_11009
            table.append(row)

        df = DataFrame(table)
        # Sort dataframe columns
        df = df.reindex(labels=['Nr.',
                                'Themenschwerpunkt',
                                'Risikobezeichnung',
                                'Beschreibung',
                                'Bearbeitungsstand',
                                'Risikoart',
                                'Erstelldatum',
                                'Risikoeinstufung',
                                'Verantwortung im Projekt',
                                'Verantwortung der Umsetzung',
                               ],
                        axis=1)
        # Clean up Status categories
        risklevel_mapping = {'High': 'Hoch',
                             'Medium': 'Mittel',
                             'Low':'Gering'}
        df['Risikoeinstufung'] = df.Risikoeinstufung.transform(lambda x: risklevel_mapping.get(x,x))
        return df

    def create_excel(self, fileobject):

        self.fileobject = fileobject
        self.opl = self.generate_opl()
        self.risk = self.generate_risk()

        # Define an Excel document
        book = load_workbook('template/Template.xlsx')
        writer = ExcelWriter(self.fileobject, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        self.opl.to_excel(writer, "Offene Punkte", startrow=6, index=False, header=False)

        # Formatting
        ws = writer.sheets['Offene Punkte']

        # Row height adjustment
        max_rows = self.opl.apply(lambda x: max([len(str(row).split()) for row in x if row is not None]), axis=1).tolist()
        for row,i in enumerate(max_rows):
            ws.row_dimensions[row+7].height = i*2.25

        # Format date columns
        date_cols = ws['E7:101']
        date_cols += ws['F7:101']
        for cell in date_cols:
            cell[0].number_format = 'DD.MM.YY'

        # Formatting
        ws = writer.sheets['Projektrisiken']

            # Row height adjustment
        max_rows = self.risk.apply(lambda x: max([len(str(row).split()) for row in x if row is not None]), axis=1).tolist()
        for row,i in enumerate(max_rows):
            ws.row_dimensions[row+7].height = i*2.25

        self.risk.to_excel(writer, 'Projektrisiken', startrow=6, index=False, header=False)

        # Save result to Excel document
        writer.save()
